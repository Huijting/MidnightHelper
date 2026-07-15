#!/usr/bin/env python3
"""Build the self-contained keybind cheat-sheet HTML (for the Artifact) and the
XLSX export, from keybinds.json (produced by gen_keybinds.py)."""
import json, os, html
HERE = os.path.dirname(os.path.abspath(__file__))
DATA = json.load(open(os.path.join(HERE, "keybinds.json"), encoding="utf-8"))
SPECS = DATA["specs"]

# label -> role group (color/legend). Keep in sync with ROLE_LABEL/CAT_LABEL.
ROLE_GROUP = {
    "Interrupt":"interrupt","Movement":"move","Mobility":"move","Utility":"util",
    "Defensive":"def","Defensive (major)":"def","Dispel / CC":"dispel","Big cooldown":"cd",
    "Cooldown":"cd","Self-heal":"heal","Heal (out of combat)":"heal","Sustain / HoT":"heal",
    "Raid heal":"heal","Rotation":"rot","Spender":"spend","Taunt":"taunt","Pet":"util","Blessing":"util",
}
GROUP_LABEL = {"interrupt":"Interrupt","move":"Movement","def":"Defensive","dispel":"Dispel / CC",
    "cd":"Cooldown","heal":"Heal","rot":"Rotation","spend":"Spender","util":"Utility","taunt":"Taunt"}

# ---------------------------------------------------------------- HTML ----
KB_ROWS = [  # physical rows of the left-hand combat zone we render
    ["F1","F2","F3","F4"],
    ["1","2","3","4","5"],
    ["Q","W","E","R","T"],
    ["A","S","D","F","G"],
    ["Z","X","C","V"],
]
MOVE_KEYS = {"W","A","S","D"}  # shown as movement, greyed

def build_html():
    payload = json.dumps({"specs":SPECS,"roleGroup":ROLE_GROUP,"groupLabel":GROUP_LABEL,"rows":KB_ROWS,
                          "moveKeys":list(MOVE_KEYS)}, ensure_ascii=False)
    # NB: content only — the Artifact wrapper supplies <!doctype>/<head>/<body>.
    return TEMPLATE.replace("/*DATA*/", payload)

TEMPLATE = r"""<title>Midnight Helper — Keybind Cheat Sheet</title>
<style>
:root{
  --bg:#f5f4f9; --panel:#ffffff; --panel2:#efedf5; --ink:#1b1b26; --muted:#5c5b70; --faint:#8d8ca2;
  --line:#dedce8; --accent:#b07d22; --accent-ink:#7a5514;
  --key-bg:#edeaf3; --key-line:#cfccdd; --key-empty:#f3f1f8;
  --r-interrupt:#cf4740; --r-move:#2c968f; --r-def:#4a7ec0; --r-dispel:#8f5fc6;
  --r-cd:#c0851f; --r-heal:#3b9a5f; --r-rot:#75748c; --r-spend:#cb7534; --r-util:#7c7b92; --r-taunt:#c46435;
}
@media (prefers-color-scheme:dark){:root{
  --bg:#0e0e17; --panel:#16161f; --panel2:#1c1c2a; --ink:#eae9f2; --muted:#a3a2b8; --faint:#75748c;
  --line:#2a2a3b; --accent:#e2b45c; --accent-ink:#f0cd86;
  --key-bg:#1e1e2c; --key-line:#35354a; --key-empty:#161622;
  --r-interrupt:#e5645c; --r-move:#41b3ab; --r-def:#6a9ede; --r-dispel:#ab7de0;
  --r-cd:#e0ad4b; --r-heal:#54b97a; --r-rot:#9695ad; --r-spend:#e08f52; --r-util:#9897ae; --r-taunt:#e08152;
}}
:root[data-theme="light"]{
  --bg:#f5f4f9; --panel:#ffffff; --panel2:#efedf5; --ink:#1b1b26; --muted:#5c5b70; --faint:#8d8ca2;
  --line:#dedce8; --accent:#b07d22; --accent-ink:#7a5514;
  --key-bg:#edeaf3; --key-line:#cfccdd; --key-empty:#f3f1f8;
  --r-interrupt:#cf4740; --r-move:#2c968f; --r-def:#4a7ec0; --r-dispel:#8f5fc6;
  --r-cd:#c0851f; --r-heal:#3b9a5f; --r-rot:#75748c; --r-spend:#cb7534; --r-util:#7c7b92; --r-taunt:#c46435;
}
:root[data-theme="dark"]{
  --bg:#0e0e17; --panel:#16161f; --panel2:#1c1c2a; --ink:#eae9f2; --muted:#a3a2b8; --faint:#75748c;
  --line:#2a2a3b; --accent:#e2b45c; --accent-ink:#f0cd86;
  --key-bg:#1e1e2c; --key-line:#35354a; --key-empty:#161622;
  --r-interrupt:#e5645c; --r-move:#41b3ab; --r-def:#6a9ede; --r-dispel:#ab7de0;
  --r-cd:#e0ad4b; --r-heal:#54b97a; --r-rot:#9695ad; --r-spend:#e08f52; --r-util:#9897ae; --r-taunt:#e08152;
}
*{box-sizing:border-box}
.sr-only{position:absolute;width:1px;height:1px;padding:0;margin:-1px;overflow:hidden;clip:rect(0 0 0 0);border:0}
body,.wrap{margin:0}
.wrap{max-width:1040px;margin:0 auto;padding:22px 20px 60px;color:var(--ink);
  font-family:ui-sans-serif,system-ui,-apple-system,"Segoe UI",Roboto,sans-serif;line-height:1.5}
h1{font-size:26px;letter-spacing:-.01em;margin:0;text-wrap:balance;font-weight:750}
.sub{color:var(--muted);margin:4px 0 0;font-size:14px}
.kbd{font-family:ui-monospace,"SF Mono",Menlo,Consolas,monospace;font-variant-numeric:tabular-nums}
.eyebrow{font-size:11px;letter-spacing:.14em;text-transform:uppercase;color:var(--accent-ink);font-weight:650}

.controls{margin-top:20px;display:flex;flex-direction:column;gap:12px}
.tabs{display:flex;flex-wrap:wrap;gap:6px}
.tab{border:1px solid var(--line);background:var(--panel);color:var(--muted);border-radius:999px;
  padding:5px 12px;font-size:13px;cursor:pointer;font-weight:600}
.tab[aria-selected="true"]{background:var(--accent);border-color:var(--accent);color:#fff}
.tab:focus-visible{outline:2px solid var(--accent);outline-offset:2px}
.specs{display:flex;flex-wrap:wrap;gap:6px;min-height:34px}
.spec{border:1px solid var(--line);background:var(--panel2);color:var(--ink);border-radius:8px;
  padding:6px 11px;font-size:13px;cursor:pointer}
.spec[aria-selected="true"]{border-color:var(--accent);box-shadow:inset 0 0 0 1px var(--accent);font-weight:650}
.spec .role{font-size:10px;color:var(--faint);text-transform:uppercase;letter-spacing:.08em;margin-left:6px}
.spec:focus-visible{outline:2px solid var(--accent);outline-offset:2px}

.bar{display:flex;flex-wrap:wrap;align-items:center;gap:14px;justify-content:space-between}
.layers{display:inline-flex;border:1px solid var(--line);border-radius:8px;overflow:hidden}
.layers button{border:0;background:var(--panel);color:var(--muted);padding:6px 12px;font-size:13px;cursor:pointer;
  border-right:1px solid var(--line);font-weight:600}
.layers button:last-child{border-right:0}
.layers button[aria-selected="true"]{background:var(--panel2);color:var(--ink)}
.layers button:focus-visible{outline:2px solid var(--accent);outline-offset:-2px}
.legend{display:flex;flex-wrap:wrap;gap:10px 14px;font-size:12px;color:var(--muted)}
.legend span{display:inline-flex;align-items:center;gap:5px}
.dot{width:9px;height:9px;border-radius:3px;display:inline-block}

.board{margin-top:18px;background:var(--panel);border:1px solid var(--line);border-radius:16px;
  padding:18px;overflow-x:auto}
.now{display:flex;align-items:baseline;gap:10px;flex-wrap:wrap;margin-bottom:14px}
.now h2{font-size:17px;margin:0;font-weight:700}
.now .layerlab{font-size:12px;color:var(--faint)}
.rows{display:flex;flex-direction:column;gap:8px;min-width:560px}
.row{display:flex;gap:8px}
.key{width:104px;min-height:66px;border:1px solid var(--key-line);border-radius:10px;background:var(--key-empty);
  padding:7px 8px;display:flex;flex-direction:column;gap:3px;position:relative}
.key .cap{font-family:ui-monospace,monospace;font-size:11px;color:var(--faint);font-weight:600}
.key .ab{font-size:12px;line-height:1.25;font-weight:600;color:var(--ink)}
.key .rl{font-size:9.5px;text-transform:uppercase;letter-spacing:.05em;color:var(--faint);margin-top:auto}
.key.filled{background:var(--key-bg)}
.key.filled{border-left:4px solid var(--rc,var(--key-line))}
.key .ab{color:var(--rc,var(--ink))}
.key.move{background:var(--panel2)}
.key.move .ab{color:var(--muted);font-weight:600}
.key.spacer{border:0;background:transparent}
.mouse{display:flex;gap:8px;margin-top:8px;align-items:center}
.mouse .cap{color:var(--faint);font-size:11px}

.info{margin-top:22px;display:grid;grid-template-columns:1fr 1fr;gap:16px}
@media (max-width:720px){.info{grid-template-columns:1fr}.rows{min-width:520px}}
.card{background:var(--panel);border:1px solid var(--line);border-radius:14px;padding:16px}
.card h3{margin:0 0 8px;font-size:14px;font-weight:700}
.card p,.card li{font-size:13px;color:var(--muted);margin:6px 0}
.card ul{margin:6px 0 0;padding-left:18px}
.big{font-size:30px;font-weight:750;color:var(--accent-ink);line-height:1}
.cc{font-size:12px;color:var(--muted)}
.cc b{color:var(--ink)}
.foot{margin-top:26px;font-size:11.5px;color:var(--faint);text-align:center;line-height:1.6}
.situ{margin-top:12px;font-size:12.5px;color:var(--muted);background:var(--panel2);
  border:1px solid var(--line);border-radius:12px;padding:11px 14px;line-height:1.6}
.situ b{color:var(--ink)}
</style>

<div class="wrap">
  <h2 class="sr-only">Interactive keyboard cheat sheet: pick a class and spec to see which World of Warcraft ability sits on each key, following Midnight Helper's v6 keybind standard.</h2>
  <div class="eyebrow">Midnight Helper</div>
  <h1>Keybind cheat sheet</h1>
  <p class="sub">One consistent key layout for every spec — same key, same kind of ability, so muscle memory carries between alts. Pick your spec, then set these in WoW.</p>

  <div class="controls">
    <div class="tabs" id="tabs" role="tablist" aria-label="Class"></div>
    <div class="specs" id="specList" aria-label="Spec"></div>
    <div class="bar">
      <div class="layers" id="layers" role="tablist" aria-label="Modifier layer"></div>
      <div class="legend" id="legend"></div>
    </div>
  </div>

  <section class="board" aria-live="polite">
    <div class="now"><h2 id="nowSpec"></h2><span class="layerlab" id="nowLayer"></span></div>
    <div class="rows" id="rows"></div>
    <div class="mouse"><span class="cap kbd">Mouse:</span>
      <span class="cc">Button 4/5 → movement + trinket · single-target heals go on <b>mouseover / click-cast</b> over the raid frames (Midnight-native), not on keys.</span>
    </div>
  </section>
  <div class="situ" id="situ"></div>

  <div class="info">
    <div class="card">
      <h3>Action bars in WoW</h3>
      <p>This spec uses <span class="big" id="bindCount">0</span> keybinds across the base + Shift / Ctrl / Alt layers.</p>
      <ul>
        <li>Open <b>Options → Action Bars</b> (or Edit Mode) and enable enough bars for every ability to have a visible slot — usually the main bar plus <b id="barCount">2</b> extra.</li>
        <li>Bind the base keys (1–5, Q E R F T X Z C V, F1–F4) on your main bar via <b>Keybindings</b>.</li>
        <li>Put the Shift / Ctrl / Alt layers on the extra bars and let bar paging swap them — same key, one modifier deeper.</li>
      </ul>
    </div>
    <div class="card">
      <h3>The anchors (never move)</h3>
      <p>These sit on the same key on every spec, so the reflex is identical whatever you play:</p>
      <ul>
        <li><span class="kbd">E</span> interrupt · <span class="kbd">Q</span> movement · <span class="kbd">Z</span>/<span class="kbd">C</span> defensives · <span class="kbd">V</span> dispel/CC</li>
        <li><span class="kbd">F1</span> big cooldown · <span class="kbd">F2–F4</span> heals · <span class="kbd">1–5</span> your rotation, <span class="kbd">Shift+</span> = the AoE twin</li>
      </ul>
      <p class="cc" id="ccNote"></p>
    </div>
  </div>

  <p class="foot">Generated from Midnight Helper's v6 keybind standard — the same layout the in-game coach shows. Spell names are the English client names. Talent choices vary, so a few keys may differ from your exact build.</p>
</div>

<script>
const D = /*DATA*/;
const specs = D.specs, roleGroup = D.roleGroup, groupLabel = D.groupLabel, rows = D.rows, moveKeys = new Set(D.moveKeys);
const LAYERS = [["","Base"],["Shift+","Shift"],["Ctrl+","Ctrl"],["Alt+","Alt"]];
let curSpec = specs.find(s=>s.id===66) || specs[0];  // default: Prot Paladin
let curLayer = "";

const byClass = {};
specs.forEach(s=>{ (byClass[s.class]=byClass[s.class]||[]).push(s); });
const classes = Object.keys(byClass).sort();

function groupOf(label){ return roleGroup[label] || "rot"; }
function color(group){ return getComputedStyle(document.documentElement).getPropertyValue('--r-'+group) || 'var(--rot)'; }

function renderTabs(){
  const t = document.getElementById('tabs'); t.innerHTML='';
  classes.forEach(c=>{
    const b=document.createElement('button'); b.className='tab'; b.textContent=c; b.setAttribute('role','tab');
    b.setAttribute('aria-selected', curSpec.class===c);
    b.onclick=()=>{ curSpec = byClass[c][0]; render(); };
    t.appendChild(b);
  });
}
function renderSpecs(){
  const el=document.getElementById('specList'); el.innerHTML='';
  byClass[curSpec.class].forEach(s=>{
    const b=document.createElement('button'); b.className='spec'; b.setAttribute('aria-selected', s.id===curSpec.id);
    b.innerHTML = s.spec + '<span class="role">'+s.role+'</span>';
    b.onclick=()=>{ curSpec=s; render(); };
    el.appendChild(b);
  });
}
function renderLayers(){
  const el=document.getElementById('layers'); el.innerHTML='';
  LAYERS.forEach(([pfx,lab])=>{
    const b=document.createElement('button'); b.textContent=lab; b.setAttribute('role','tab');
    b.setAttribute('aria-selected', curLayer===pfx);
    b.onclick=()=>{ curLayer=pfx; render(); };
    el.appendChild(b);
  });
}
function renderLegend(){
  const el=document.getElementById('legend'); el.innerHTML='';
  ["interrupt","move","def","dispel","cd","heal","rot","spend","util","taunt"].forEach(g=>{
    const s=document.createElement('span');
    s.innerHTML='<span class="dot" style="background:var(--r-'+g+')"></span>'+groupLabel[g];
    el.appendChild(s);
  });
}
function keyFor(base){ return curSpec.keys[curLayer+base]; }
function renderBoard(){
  document.getElementById('nowSpec').textContent = curSpec.class+' — '+curSpec.spec;
  document.getElementById('nowLayer').textContent = curLayer ? (curLayer.replace('+','')+' layer') : 'Base keys';
  const host=document.getElementById('rows'); host.innerHTML='';
  rows.forEach(r=>{
    const rowEl=document.createElement('div'); rowEl.className='row';
    r.forEach(base=>{
      const k=document.createElement('div'); k.className='key';
      const cap=document.createElement('div'); cap.className='cap'; cap.textContent=(curLayer||'')+base; k.appendChild(cap);
      const hit = keyFor(base);
      if(moveKeys.has(base) && !hit){
        k.classList.add('move');
        const ab=document.createElement('div'); ab.className='ab'; ab.textContent='Move'; k.appendChild(ab);
      } else if(hit){
        k.classList.add('filled');
        const g=groupOf(hit.label); k.style.setProperty('--rc','var(--r-'+g+')');
        const ab=document.createElement('div'); ab.className='ab'; ab.textContent=hit.name; k.appendChild(ab);
        const rl=document.createElement('div'); rl.className='rl'; rl.textContent=hit.label||''; k.appendChild(rl);
        k.title = hit.name+' — '+(hit.label||'');
      }
      rowEl.appendChild(k);
    });
    host.appendChild(rowEl);
  });
}
function renderInfo(){
  const n = Object.keys(curSpec.keys).length;
  document.getElementById('bindCount').textContent = n;
  const distinctBase = new Set(Object.keys(curSpec.keys).map(k=>k.split('+').pop()));
  const bars = Math.max(2, Math.ceil(n/12));
  document.getElementById('barCount').textContent = (bars-1);
  const cc = curSpec.clickcast||[];
  document.getElementById('ccNote').textContent = cc.length
     ? ('Click-cast / mouseover (not on keys): '+cc.join(', ')+'.') : '';
  const situ=document.getElementById('situ'); const sl=curSpec.situational||[];
  if(sl.length){ situ.style.display='';
    situ.innerHTML='<b>Situational</b> — extra CC / dispels with no fixed home key; bind these where you like: '+sl.map(x=>x.name).join(', ')+'.';
  } else { situ.style.display='none'; }
}
function render(){ renderTabs(); renderSpecs(); renderLayers(); renderLegend(); renderBoard(); renderInfo(); }
render();
</script>
"""

def build_xlsx():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()
    thin = Side(style="thin", color="D0CEDC")
    border = Border(left=thin,right=thin,top=thin,bottom=thin)
    GRP_HEX = {"interrupt":"CF4740","move":"2C968F","def":"4A7EC0","dispel":"8F5FC6","cd":"C0851F",
        "heal":"3B9A5F","rot":"75748C","spend":"CB7534","util":"7C7B92","taunt":"C46435"}
    def hdr(ws, cells):
        for i,(t) in enumerate(cells,1):
            c=ws.cell(row=1,column=i,value=t); c.font=Font(bold=True,color="FFFFFF")
            c.fill=PatternFill("solid",fgColor="2E2E44"); c.border=border
    # ---- Keyboard sheet (the v6 layout as a grid) ----
    ws=wb.active; ws.title="Keyboard (v6)"
    ws["A1"]="Midnight Helper — v6 keybind layout (roles per key). See the per-spec tabs for the abilities."
    ws["A1"].font=Font(bold=True,size=13)
    layout=[["F1","F2","F3","F4"],["1","2","3","4","5"],["Q","W","E","R","T"],["A","S","D","F","G"],["Z","X","C","V"]]
    ROLEKEY={"E":"Interrupt","Q":"Movement","Z":"Defensive","C":"Defensive (major)","V":"Dispel / CC",
      "F1":"Big cooldown","F2":"Self-heal","F3":"Heal (OOC)","F4":"Sustain / HoT",
      "1":"Rotation","2":"Rotation","3":"Rotation","4":"Spender","5":"Spender",
      "R":"Utility/Mobility","F":"Utility/Taunt","X":"Defensive","T":"Utility","W":"Move","A":"Move","S":"Move","D":"Move"}
    r0=3
    for ri,row in enumerate(layout):
        for ci,k in enumerate(row):
            cell=ws.cell(row=r0+ri*2, column=1+ci, value=k)
            cell.font=Font(bold=True); cell.alignment=Alignment(horizontal="center")
            cell.border=border; cell.fill=PatternFill("solid",fgColor="EDEAF3")
            rl=ws.cell(row=r0+ri*2+1, column=1+ci, value=ROLEKEY.get(k,""))
            rl.font=Font(size=9,color="666666"); rl.alignment=Alignment(horizontal="center"); rl.border=border
    for c in range(1,6): ws.column_dimensions[get_column_letter(c)].width=16
    note=r0+len(layout)*2+2
    ws.cell(row=note,column=1,value="Overflow: when a group is full, the same key moves one modifier deeper — Shift → Ctrl → Alt.").font=Font(italic=True,size=10)
    ws.cell(row=note+1,column=1,value="AoE = the Shift-twin of its single-target key (1 → Shift+1, 4 → Shift+4).").font=Font(italic=True,size=10)
    # ---- per-spec sheets ----
    for s in SPECS:
        title=(s["class"][:14]+" "+s["spec"])[:31]
        ws=wb.create_sheet(title=title)
        hdr(ws,["Key","Ability","Role"])
        # order rows: base layer keys in reading order, then modifier layers
        order=["E","Q","1","2","3","4","5","F","R","X","T","Z","C","V","F1","F2","F3","F4"]
        def sortkey(bk):
            base=bk.split("+")[-1]; mod=bk.split("+")[0] if "+" in bk else ""
            mo={"":0,"Shift":1,"Ctrl":2,"Alt":3}.get(mod,4)
            bi=order.index(base) if base in order else 99
            return (mo,bi,base)
        rown=2
        for bk in sorted(s["keys"].keys(), key=sortkey):
            info=s["keys"][bk]; g=ROLE_GROUP.get(info["label"],"rot")
            ws.cell(row=rown,column=1,value=bk).font=Font(bold=True,name="Consolas")
            ws.cell(row=rown,column=2,value=info["name"])
            rc=ws.cell(row=rown,column=3,value=info["label"]); rc.font=Font(color=GRP_HEX.get(g,"75748C"),bold=True)
            for cc in range(1,4): ws.cell(row=rown,column=cc).border=border
            rown+=1
        if s.get("situational"):
            ws.cell(row=rown+1,column=1,value="Situational (no fixed key — bind where you like):").font=Font(italic=True)
            ws.cell(row=rown+1,column=2,value=", ".join(x["name"] for x in s["situational"]))
            rown+=2
        if s.get("clickcast"):
            ws.cell(row=rown+1,column=1,value="Click-cast / mouseover (raid frames, not keys):").font=Font(italic=True)
            ws.cell(row=rown+1,column=2,value=", ".join(s["clickcast"]))
        ws.column_dimensions["A"].width=12; ws.column_dimensions["B"].width=30; ws.column_dimensions["C"].width=18
    outp=os.path.join(HERE,"MidnightHelper_Keybinds.xlsx"); wb.save(outp); return outp

html_out=os.path.join(HERE,"keybind_cheatsheet.html")
open(html_out,"w",encoding="utf-8").write(build_html())
print("HTML:", html_out, os.path.getsize(html_out),"bytes")
print("XLSX:", build_xlsx())
